#!/usr/bin/env python3
"""
Investor Portal with Authentication
Secure login for Eric and Phillip to view their individual investment data
"""

from flask import Flask, request, jsonify, render_template, redirect, url_for, session, flash
from flask_cors import CORS
from functools import wraps
import os
from dotenv import load_dotenv
from database import DatabaseManager
from auth import AuthManager
from chatbot import ChatbotManager
import logging
from datetime import datetime
import pandas as pd

# Load environment variables
load_dotenv()

def clean_sql_data(text):
    """Clean text data to prevent SQL injection and syntax errors"""
    if not text:
        return ""
    
    # Convert to string if not already
    text = str(text)
    
    # Remove or replace problematic characters
    text = text.replace("'", "''")  # Escape single quotes
    text = text.replace('"', '""')  # Escape double quotes
    text = text.replace('\n', ' ')  # Replace newlines with spaces
    text = text.replace('\r', ' ')  # Replace carriage returns with spaces
    text = text.replace('\t', ' ')  # Replace tabs with spaces
    
    # Remove any remaining control characters
    import re
    text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
    
    # Limit length to prevent issues
    if len(text) > 1000:
        text = text[:1000] + "..."
    
    return text.strip()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# Get port from environment variable (for deployment platforms)
PORT = int(os.environ.get('PORT', 5002))
CORS(app)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize components
# Use persistent disk for database in production (Render uses /data for persistent disks)
# Fallback to local path for development
DB_PATH = os.environ.get('DATABASE_PATH', 
                        '/data/insurance_crm.db' if os.path.exists('/data') else 'insurance_crm.db')
db_manager = DatabaseManager(DB_PATH)
auth_manager = AuthManager(DB_PATH)

# Initialize chatbot manager (optional - only if OPENAI_API_KEY is set)
try:
    chatbot_manager = ChatbotManager(DB_PATH)
except ValueError as e:
    # OPENAI_API_KEY not set - chatbot features disabled
    logger.warning(f"Chatbot disabled: {e}")
    chatbot_manager = None
except Exception as e:
    logger.warning(f"Failed to initialize chatbot: {e}")
    chatbot_manager = None

# Auto-setup on startup
def auto_setup_on_startup():
    """Run auto-setup if needed when the app starts"""
    try:
        # Check if investors exist
        investors = db_manager.get_investors()
        if len(investors) == 0:
            print("🚀 Auto-setup: Creating investors...")
            db_manager.add_investor('Eric', 42.0)
            db_manager.add_investor('Phillip', 40.0)
            print("✅ Auto-setup: Investors created")
        
        # Check if users exist by trying to verify them
        eric_check = auth_manager.verify_user('eric', 'eric123')
        admin_check = auth_manager.verify_user('admin', 'admin123')
        
        if not eric_check['success'] or not admin_check['success']:
            print("🚀 Auto-setup: Creating user accounts...")
            
            # Get investor IDs
            investors = db_manager.get_investors()
            eric_investor = next((i for i in investors if i['name'].lower() == 'eric'), None)
            phillip_investor = next((i for i in investors if i['name'].lower() == 'phillip'), None)
            
            # Create users with proper parameters
            auth_manager.create_user('eric', 'eric123', eric_investor['id'] if eric_investor else None, False)
            auth_manager.create_user('phillip', 'phillip123', phillip_investor['id'] if phillip_investor else None, False)
            auth_manager.create_user('admin', 'admin123', None, True)
            print("✅ Auto-setup: User accounts created")
        
        # Check if we have any data, if not, try to import from uploads directory
        leads = db_manager.get_leads()
        enrollments = db_manager.get_enrollments()
        
        if len(leads) == 0 and len(enrollments) == 0:
            print("📊 Auto-setup: No data found, checking for CSV files...")
            auto_import_csv_data()
            
    except Exception as e:
        print(f"⚠️ Auto-setup warning: {e}")

def auto_import_csv_data():
    """Automatically import CSV data if files exist in uploads directory"""
    try:
        # Check for uploads directory - use persistent disk if available
        if os.path.exists('/data'):
            uploads_dir = os.getenv('UPLOADS_PATH', '/data/uploads')
        else:
            uploads_dir = os.getenv('UPLOADS_PATH', 'uploads')
        if not os.path.exists(uploads_dir):
            print("📁 Auto-setup: No uploads directory found")
            return
        
        # Look for CSV files
        csv_files = []
        for file in os.listdir(uploads_dir):
            if file.endswith('.csv'):
                csv_files.append(file)
        
        if not csv_files:
            print("📁 Auto-setup: No CSV files found in uploads directory")
            return
        
        print(f"📁 Auto-setup: Found {len(csv_files)} CSV files")
        
        # Import each CSV file
        for csv_file in csv_files:
            file_path = os.path.join(uploads_dir, csv_file)
            print(f"📊 Auto-setup: Importing {csv_file}...")
            
            # Determine investor and data type from filename
            investor_name = None
            data_type = None
            
            if 'eric' in csv_file.lower():
                investor_name = 'Eric'
            elif 'phill' in csv_file.lower():
                investor_name = 'Phillip'
            
            if 'lead' in csv_file.lower():
                data_type = 'leads'
            elif 'enrollment' in csv_file.lower():
                data_type = 'enrollments'
            
            if investor_name and data_type:
                # Import the data
                import_csv_data(file_path, investor_name, data_type)
                print(f"✅ Auto-setup: Imported {csv_file}")
            else:
                print(f"⚠️ Auto-setup: Could not determine investor/type for {csv_file}")
                
    except Exception as e:
        print(f"⚠️ Auto-setup CSV import warning: {e}")

def import_csv_data(file_path, investor_name, data_type):
    """Import CSV data (simplified version of the web upload logic)"""
    try:
        import pandas as pd
        from datetime import datetime
        import re
        
        # Read CSV
        df = pd.read_csv(file_path)
        rows_data = df.to_dict(orient='records')
        
        # Get investor
        investors = db_manager.get_investors()
        investor = next((i for i in investors if i['name'].lower() == investor_name.lower()), None)
        
        if not investor:
            print(f"❌ Auto-setup: Investor {investor_name} not found")
            return
        
        records_imported = 0
        
        for row in rows_data:
            # Extract and clean data
            first_name = str(row.get('First Name', '')).strip() if pd.notna(row.get('First Name', '')) else ''
            last_name = str(row.get('Last Name', '')).strip() if pd.notna(row.get('Last Name', '')) else ''
            insured_name = f"{first_name} {last_name}".strip()
            
            if not insured_name or insured_name.strip() == "" or insured_name == " ":
                continue
            
            # Clean the name
            insured_name = re.sub(r'[^\w\s\-\.]', '', insured_name)
            insured_name = insured_name.strip()
            insured_name = clean_sql_data(insured_name)
            
            if not insured_name or insured_name.strip() == "":
                continue
            
            # Parse date
            try:
                created_date = datetime.fromisoformat(row.get('Created', '').replace('Z', '+00:00')).strftime('%Y-%m-%d')
            except:
                created_date = datetime.now().strftime('%Y-%m-%d')
            
            # Clean notes
            tags_value = str(row.get('Tags', '')) if pd.notna(row.get('Tags', '')) else ''
            notes = clean_sql_data(tags_value)
            if notes:
                notes = f"Tags: {notes}"
            else:
                notes = "Auto-imported from CSV"
            
            # Import based on data type
            if data_type == 'leads':
                result = db_manager.add_lead(
                    investor_id=investor['id'],
                    insured_name=insured_name,
                    lead_date=created_date,
                    notes=notes
                )
            elif data_type == 'enrollments':
                result = db_manager.add_enrollment(
                    insured_name=insured_name,
                    enrollment_date=created_date,
                    labor_cost=15.00,
                    notes=notes
                )
            
            if result['success']:
                records_imported += 1
        
        print(f"📊 Auto-setup: Imported {records_imported} records from {os.path.basename(file_path)}")
        
    except Exception as e:
        print(f"❌ Auto-setup CSV import error: {e}")

# Run auto-setup
auto_setup_on_startup()

# ============== AUTHENTICATION DECORATORS ==============

def login_required(f):
    """Decorator to require login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def investor_only(f):
    """Decorator to require investor role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        
        user = auth_manager.get_user_by_id(session['user_id'])
        if not user or not user.get('investor_id'):
            flash('Access denied. Investors only.', 'error')
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

def admin_only(f):
    """Decorator to require admin role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        
        user = auth_manager.get_user_by_id(session['user_id'])
        if not user or not user.get('is_admin'):
            flash('Access denied. Admin only.', 'error')
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

# ============== AUTHENTICATION ROUTES ==============

@app.route('/')
def index():
    """Landing page"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        result = auth_manager.verify_user(username, password)
        
        if result['success']:
            user = result['user']
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['investor_id'] = user.get('investor_id')
            session['investor_name'] = user.get('investor_name')
            session['is_admin'] = user.get('is_admin', False)
            
            flash(f'Welcome back, {user.get("investor_name") or username}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash(result.get('error', 'Invalid credentials'), 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Logout"""
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

# ============== DASHBOARD ROUTES ==============

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard - shows investor-specific data"""
    user = auth_manager.get_user_by_id(session['user_id'])
    
    if not user or not user.get('investor_id'):
        # Admin view - show all investors' data
        contributions = db_manager.get_investor_contributions()
        cost_analysis = db_manager.get_cost_analysis()
        
        admin_data = {
            'user': user,
            'contributions': contributions,
            'cost_analysis': cost_analysis
        }
        return render_template('admin.html', **admin_data)
    
    # Investor view - show only their data
    investor_id = user['investor_id']
    
    # Get investor contributions
    contributions = db_manager.get_investor_contributions(investor_id)
    
    # Get investor's leads
    leads = db_manager.get_leads(investor_id=investor_id)
    
    # Get investor's enrollments (through leads)
    enrollments = db_manager.get_enrollments()
    investor_enrollments = [e for e in enrollments if e.get('investor_id') == investor_id]
    
    # Calculate metrics
    total_leads = len(leads)
    converted_leads = len([l for l in leads if l['status'] == 'converted'])
    active_leads = len([l for l in leads if l['status'] == 'active'])
    conversion_rate = (converted_leads / total_leads * 100) if total_leads > 0 else 0
    
    total_lead_costs = sum(l['cost'] for l in leads)
    total_labor_costs = sum(e['labor_cost'] for e in investor_enrollments)
    total_invested = total_lead_costs + total_labor_costs
    
    data = {
        'investor': user,
        'total_leads': total_leads,
        'converted_leads': converted_leads,
        'active_leads': active_leads,
        'conversion_rate': round(conversion_rate, 2),
        'total_lead_costs': total_lead_costs,
        'total_labor_costs': total_labor_costs,
        'total_invested': total_invested,
        'recent_leads': leads[:10],
        'recent_enrollments': investor_enrollments[:10]
    }
    
    return render_template('dashboard.html', **data)

# ============== API ROUTES FOR INVESTOR DATA ==============

@app.route('/api/investor/stats')
@login_required
@investor_only
def get_investor_stats():
    """Get investor statistics"""
    try:
        investor_id = session.get('investor_id')
        
        contributions = db_manager.get_investor_contributions(investor_id)
        leads = db_manager.get_leads(investor_id=investor_id)
        enrollments = db_manager.get_enrollments()
        
        investor_enrollments = [e for e in enrollments if e.get('investor_id') == investor_id]
        
        stats = {
            'success': True,
            'contributions': contributions,
            'total_leads': len(leads),
            'active_leads': len([l for l in leads if l['status'] == 'active']),
            'converted_leads': len([l for l in leads if l['status'] == 'converted']),
            'total_enrollments': len(investor_enrollments),
            'total_labor_costs': sum(e['labor_cost'] for e in investor_enrollments)
        }
        
        return jsonify(stats)
        
    except Exception as e:
        logger.error(f"Error getting investor stats: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/investor/leads')
@login_required
@investor_only
def get_investor_leads():
    """Get investor's leads"""
    try:
        investor_id = session.get('investor_id')
        status = request.args.get('status')
        
        leads = db_manager.get_leads(investor_id=investor_id, status=status)
        
        return jsonify({'success': True, 'leads': leads})
        
    except Exception as e:
        logger.error(f"Error getting investor leads: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/investor/enrollments')
@login_required
@investor_only
def get_investor_enrollments():
    """Get investor's enrollments"""
    try:
        investor_id = session.get('investor_id')
        
        enrollments = db_manager.get_enrollments()
        investor_enrollments = [e for e in enrollments if e.get('investor_id') == investor_id]
        
        return jsonify({'success': True, 'enrollments': investor_enrollments})
        
    except Exception as e:
        logger.error(f"Error getting investor enrollments: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/investor/roi')
@login_required
@investor_only
def get_investor_roi():
    """Calculate investor ROI"""
    try:
        investor_id = session.get('investor_id')
        
        # Get leads and calculate costs
        leads = db_manager.get_leads(investor_id=investor_id)
        total_lead_costs = sum(l['cost'] for l in leads)
        
        # Get enrollments and labor costs
        enrollments = db_manager.get_enrollments()
        investor_enrollments = [e for e in enrollments if e.get('investor_id') == investor_id]
        total_labor_costs = sum(e['labor_cost'] for e in investor_enrollments)
        
        # Total invested
        total_invested = total_lead_costs + total_labor_costs
        
        # Get commission data (if available)
        # This would need to be linked to specific enrollments
        total_commission = 0  # Placeholder
        
        # Calculate ROI
        net_profit = total_commission - total_invested
        roi_percentage = (net_profit / total_invested * 100) if total_invested > 0 else 0
        
        return jsonify({
            'success': True,
            'total_invested': round(total_invested, 2),
            'total_lead_costs': round(total_lead_costs, 2),
            'total_labor_costs': round(total_labor_costs, 2),
            'total_commission': round(total_commission, 2),
            'net_profit': round(net_profit, 2),
            'roi_percentage': round(roi_percentage, 2)
        })
        
    except Exception as e:
        logger.error(f"Error calculating investor ROI: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/upload-investor-data', methods=['POST'])
@login_required
def admin_upload_investor_data():
    """Admin endpoint to upload and import investor CSV data"""
    try:
        # Verify admin access
        user = auth_manager.get_user_by_id(session['user_id'])
        if not user or not user.get('is_admin'):
            return jsonify({'success': False, 'error': 'Admin access required'}), 403
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        investor_name = request.form.get('investor')
        data_type = request.form.get('data_type')
        
        if not investor_name or not data_type:
            return jsonify({'success': False, 'error': 'Investor and data type required'}), 400
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        # Accept multiple file types
        allowed_extensions = ['.csv', '.xlsx', '.xls', '.pdf', '.txt']
        file_ext = os.path.splitext(file.filename)[1].lower()
        
        if file_ext not in allowed_extensions:
            return jsonify({'success': False, 'error': 'Unsupported file type. Please upload CSV, Excel, PDF, or Text files'}), 400
        
        # Save file to persistent uploads directory
        # Use /data/uploads on Render (persistent disk), or local uploads for dev
        if os.path.exists('/data'):
            # Production: Use persistent disk
            uploads_dir = os.environ.get('UPLOADS_PATH', '/data/uploads')
        else:
            # Development: Use local uploads directory
            uploads_dir = os.environ.get('UPLOADS_PATH', 'uploads')
        
        os.makedirs(uploads_dir, exist_ok=True)
        file_path = os.path.join(uploads_dir, file.filename)
        
        print(f"DEBUG: Saving file to: {file_path}")  # Debug log
        file.save(file_path)
        
        # Get investor
        investors = db_manager.get_investors()
        investor = next((i for i in investors if i['name'].lower() == investor_name.lower()), None)
        
        if not investor:
            os.remove(file_path)
            return jsonify({'success': False, 'error': f'Investor {investor_name} not found'}), 404
        
        # Import data based on file type
        import csv
        from datetime import datetime
        
        records_imported = 0
        rows_data = []
        
        # Parse file based on type
        if file_ext == '.csv':
            # Parse CSV
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows_data = list(reader)
        
        elif file_ext in ['.xlsx', '.xls']:
            # Parse Excel
            try:
                import pandas as pd
                df = pd.read_excel(file_path)
                rows_data = df.to_dict('records')
            except ImportError:
                # Fallback: try openpyxl directly
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    rows_data = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows_data.append(dict(zip(headers, row)))
                except:
                    os.remove(file_path)
                    return jsonify({'success': False, 'error': 'Excel parsing failed. Please install openpyxl or pandas'}), 500
        
        elif file_ext == '.pdf':
            # For PDF, we'll extract text and try to parse structured data
            os.remove(file_path)
            return jsonify({'success': False, 'error': 'PDF parsing not yet implemented. Please use CSV or Excel for now.'}), 400
        
        elif file_ext == '.txt':
            # Parse tab or comma-delimited text file
            with open(file_path, 'r', encoding='utf-8') as f:
                # Try to detect delimiter
                first_line = f.readline()
                delimiter = '\t' if '\t' in first_line else ','
                f.seek(0)
                reader = csv.DictReader(f, delimiter=delimiter)
                rows_data = list(reader)
        
        # Process the parsed data
        if rows_data:
            # Import pandas if available (for NaN checking)
            try:
                import pandas as pd
                has_pandas = True
            except ImportError:
                has_pandas = False
            
            reader = rows_data  # Use the parsed data as if it were from CSV reader
            
            for row in reader:
                # Extract and clean data - handle NaN values from pandas
                if has_pandas:
                    first_name = str(row.get('First Name', '')).strip() if pd.notna(row.get('First Name', '')) else ''
                    last_name = str(row.get('Last Name', '')).strip() if pd.notna(row.get('Last Name', '')) else ''
                else:
                    # Fallback when pandas not available
                    first_name_val = row.get('First Name', '')
                    last_name_val = row.get('Last Name', '')
                    first_name = str(first_name_val).strip() if first_name_val and str(first_name_val).lower() != 'nan' else ''
                    last_name = str(last_name_val).strip() if last_name_val and str(last_name_val).lower() != 'nan' else ''
                insured_name = f"{first_name} {last_name}".strip()
                
                # Clean and validate the insured_name to prevent SQL errors
                if not insured_name or insured_name.strip() == "" or insured_name == " ":
                    continue
                
                # Remove any problematic characters that could cause SQL syntax errors
                import re
                # More aggressive cleaning - remove all special characters except basic ones
                insured_name = re.sub(r'[^\w\s\-\.]', '', insured_name)  # Keep only alphanumeric, spaces, hyphens, dots
                insured_name = insured_name.strip()  # Remove leading/trailing spaces
                
                # Use our comprehensive cleaning function
                insured_name = clean_sql_data(insured_name)
                
                if not insured_name or insured_name.strip() == "":
                    print(f"DEBUG: Skipping empty name after cleaning")
                    continue
                
                print(f"DEBUG: Cleaned name: '{insured_name}'")  # Debug log
                
                try:
                    created_date = datetime.fromisoformat(row.get('Created', '').replace('Z', '+00:00')).strftime('%Y-%m-%d')
                except:
                    created_date = datetime.now().strftime('%Y-%m-%d')
                
                # Handle tags field
                if has_pandas:
                    tags = str(row.get('Tags', '')).lower() if pd.notna(row.get('Tags', '')) else ''
                else:
                    tags_val = row.get('Tags', '')
                    tags = str(tags_val).lower() if tags_val and str(tags_val).lower() != 'nan' else ''
                
                if data_type == 'leads':
                    # Import as lead
                    status = 'converted' if 'enrollment' in tags else 'active'
                    
                    # Clean the notes field to prevent SQL syntax errors
                    if has_pandas:
                        tags_value = str(row.get('Tags', '')) if pd.notna(row.get('Tags', '')) else ''
                    else:
                        tags_val = row.get('Tags', '')
                        tags_value = str(tags_val) if tags_val and str(tags_val).lower() != 'nan' else ''
                    notes = clean_sql_data(tags_value)
                    if notes:
                        notes = f"Tags: {notes}"
                    else:
                        notes = "Imported from CSV"
                    
                    result = db_manager.add_lead(
                        investor_id=investor['id'],
                        insured_name=insured_name,
                        lead_date=created_date,
                        notes=notes
                    )
                    
                    if result['success']:
                        records_imported += 1
                        # Update status if converted
                        if status == 'converted':
                            import sqlite3
                            with sqlite3.connect(db_manager.db_path) as conn:
                                cursor = conn.cursor()
                                cursor.execute('UPDATE leads SET status = ? WHERE id = ?',
                                             (status, result['lead_id']))
                                conn.commit()
                
                elif data_type == 'enrollments':
                    # Import as enrollment - process all records since this is an enrollment file for this investor
                    # For enrollment files, process all records as they belong to the selected investor
                    print(f"DEBUG: Processing enrollment for {insured_name}")  # Debug log
                    
                    # Clean the notes field to prevent SQL syntax errors
                    if has_pandas:
                        tags_value = str(row.get('Tags', '')) if pd.notna(row.get('Tags', '')) else ''
                    else:
                        tags_val = row.get('Tags', '')
                        tags_value = str(tags_val) if tags_val and str(tags_val).lower() != 'nan' else ''
                    notes = clean_sql_data(tags_value)
                    if notes:
                        notes = f"Tags: {notes}"
                    else:
                        notes = "Imported from CSV"
                    
                    try:
                        result = db_manager.add_enrollment(
                            insured_name=insured_name,
                            enrollment_date=created_date,
                            labor_cost=15.00,
                            notes=notes
                        )
                        print(f"DEBUG: Enrollment result: {result}")  # Debug log
                    except Exception as e:
                        print(f"DEBUG: Enrollment error: {e}")  # Debug log
                        logger.error(f"Error adding enrollment for {insured_name}: {e}")
                        continue
                    
                    if result['success']:
                        records_imported += 1
                        print(f"DEBUG: Successfully imported enrollment {records_imported}")  # Debug log
        
        # Clean up file
        os.remove(file_path)
        
        # Link leads to enrollments if both exist
        if data_type == 'enrollments':
            import sqlite3
            with sqlite3.connect(db_manager.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    UPDATE enrollments e
                    SET lead_id = (
                        SELECT l.id FROM leads l 
                        WHERE l.insured_name = e.insured_name 
                        AND l.enrollment_id IS NULL
                        LIMIT 1
                    )
                    WHERE e.lead_id IS NULL
                ''')
                cursor.execute('''
                    UPDATE leads
                    SET enrollment_id = (
                        SELECT e.id FROM enrollments e
                        WHERE e.insured_name = leads.insured_name
                        LIMIT 1
                    ),
                    status = 'converted'
                    WHERE enrollment_id IS NULL
                    AND EXISTS (
                        SELECT 1 FROM enrollments e
                        WHERE e.insured_name = leads.insured_name
                    )
                ''')
                conn.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully imported {data_type} for {investor_name}',
            'records_imported': records_imported
        })
        
    except Exception as e:
        logger.error(f"Error uploading investor data: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============== CHATBOT API ENDPOINTS ==============

@app.route('/api/investor/chatbot/process-data', methods=['POST'])
@login_required
@investor_only
def process_investor_data_for_chatbot():
    """Process investor's data for chatbot functionality"""
    try:
        if not chatbot_manager:
            return jsonify({
                'success': False,
                'error': 'Chatbot features are not available. Please configure OPENAI_API_KEY to enable AI-powered search.'
            }), 503
        
        user = auth_manager.get_user_by_id(session['user_id'])
        investor_id = user['investor_id']
        
        # Process both leads and enrollments data
        leads_success = chatbot_manager.process_csv_data(investor_id, 'leads')
        enrollments_success = chatbot_manager.process_csv_data(investor_id, 'enrollments')
        
        if leads_success and enrollments_success:
            return jsonify({
                'success': True, 
                'message': 'Data processed successfully for chatbot'
            })
        else:
            return jsonify({
                'success': False, 
                'error': 'Failed to process some data'
            }), 500
    
    except Exception as e:
        logger.error(f"Error processing data for chatbot: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/investor/chatbot/search', methods=['POST'])
@login_required
@investor_only
def chatbot_search():
    """Search through investor's data using chatbot"""
    try:
        if not chatbot_manager:
            return jsonify({
                'success': False,
                'error': 'Chatbot features are not available. Please configure OPENAI_API_KEY to enable AI-powered search.'
            }), 503
        
        user = auth_manager.get_user_by_id(session['user_id'])
        investor_id = user['investor_id']
        
        data = request.get_json()
        query = data.get('query', '').strip()
        data_types = data.get('data_types', ['leads', 'enrollments'])
        
        if not query:
            return jsonify({'success': False, 'error': 'Query is required'}), 400
        
        # Search through the data
        results = chatbot_manager.search_data(investor_id, query, data_types)
        
        if results:
            return jsonify({
                'success': True,
                'results': results
            })
        else:
            return jsonify({
                'success': True,
                'results': [{
                    'query': query,
                    'response': "I couldn't find specific information related to your query. Please try rephrasing your question or ask about leads, enrollments, or specific metrics.",
                    'timestamp': datetime.now().isoformat()
                }]
            })
    
    except Exception as e:
        logger.error(f"Error in chatbot search: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/investor/chatbot/history')
@login_required
@investor_only
def get_chatbot_history():
    """Get chat history for the investor"""
    try:
        if not chatbot_manager:
            return jsonify({
                'success': True,
                'history': []
            })
        
        user = auth_manager.get_user_by_id(session['user_id'])
        investor_id = user['investor_id']
        
        limit = request.args.get('limit', 10, type=int)
        history = chatbot_manager.get_chat_history(investor_id, limit)
        
        return jsonify({
            'success': True,
            'history': history
        })
    
    except Exception as e:
        logger.error(f"Error getting chat history: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/chatbot/process-all-data', methods=['POST'])
@login_required
@admin_only
def admin_process_all_chatbot_data():
    """Admin endpoint to process all investor data for chatbot"""
    try:
        if not chatbot_manager:
            return jsonify({
                'success': False,
                'error': 'Chatbot features are not available. Please configure OPENAI_API_KEY to enable AI-powered search.'
            }), 503
        
        success = chatbot_manager.process_all_investor_data()
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Successfully processed all investor data for chatbot'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to process some investor data'
            }), 500
    
    except Exception as e:
        logger.error(f"Error processing all chatbot data: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/health')
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'message': 'Investor Portal is running'})

if __name__ == '__main__':
    print("=" * 60)
    print("🚀 Starting Investor Portal")
    print("=" * 60)
    print("\n🔐 Investor Login Credentials:")
    print("   Eric:")
    print("     Username: eric")
    print("     Password: eric123")
    print("\n   Phillip:")
    print("     Username: phillip")
    print("     Password: phillip123")
    print("\n   Admin:")
    print("     Username: admin")
    print("     Password: admin123")
    print("\n🌐 Open your browser to: http://localhost:5002")
    print("=" * 60)
    
    # Use environment variable for production, or default port for local dev
    debug_mode = os.environ.get('FLASK_ENV', 'development') == 'development'
    app.run(debug=debug_mode, host='0.0.0.0', port=PORT)

